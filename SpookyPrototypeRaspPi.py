# ----------------------------------------------------- # 
    # Cornell Spooky - Tristan Kuzma
# ----------------------------------------------------- # 
#Import Libraries 
'''
Import all of the neccisary libraries. 
Go to windows powershell to download these libraries on Win.
RaspPi -> Terminal

sudo python3 /home/cornelltgod/Documents/CornellSpooky/SpookyPrototypeRaspPi.py

'''

import tkinter as tk         #pip install tk      
from tkinter import font as tkfont  
from tkinter import ttk
import PIL                   #pip install pillow
import openpyxl              #pip install openpyxl
import json 
import os 
import subprocess
from subprocess import call 
from subprocess import check_output
import time 
import threading
from threading import Thread


# ----------------------------------------------------- # 
#Load all neccisary file paths

#GUI Elements 
rightArrow = '/home/cornelltgod/Documents/CornellSpooky/GUIElements/rightArrow.png'
leftArrow = '/home/cornelltgod/Documents/CornellSpooky/GUIElements/leftArrow.png'
button = '/home/cornelltgod/Documents/CornellSpooky/GUIElements/button.png'
upload = '/home/cornelltgod/Documents/CornellSpooky/GUIElements/upload.png'
vertBar = '/home/cornelltgod/Documents/CornellSpooky/GUIElements/vertBar.png'

#Data and Code
jsonFile = '/home/cornelltgod/Documents/CornellSpooky/data.json'

serverCode = '/home/cornelltgod/Documents/CornellSpooky/pymodbus-dev/TGODSerialRWServer.py'


# ----------------------------------------------------- # 
#Load all data from the excel file. 

#Define your excel file. 
excelFile = '/home/cornelltgod/Documents/CornellSpooky/ControllerRegisters.xlsx'

#Create the dataframe. 
dataframe = openpyxl.load_workbook(excelFile)

#Set an active variable to read sheet. 
dataframe1 = dataframe.active

#Define the rows 
numRow = dataframe1.max_row
numCol = dataframe1.max_column
print("Num Col: " + str(numCol))
print("Num Row: " + str(numRow))

#Define columns 
colList = [None] * numCol
print(colList)
i = 0
for i in range(0,numCol): 
    colList[i] = dataframe1.cell(row = 1, column = i+1).value

print (colList)

ControllerRegs = [[]] * numRow
#Set the base value -> The value to load to all the registers. 
baseVal = 0

#Create nested for loop to create an array containing only slaveID == targetSlave objects. 
useRow = 1
for rowInt in range(2, numRow):
        ControllerRegs[useRow] = [(dataframe1.cell(row = rowInt, column = 1)).value, ((dataframe1.cell(row = rowInt, column = 2)).value), (dataframe1.cell(row = rowInt, column = 3)).value, (dataframe1.cell(row = rowInt, column = 4)).value, (dataframe1.cell(row = rowInt, column = 5)).value]
        print(ControllerRegs[useRow])
        print("\n")
        useRow += 1

#print(useRow)
#print(ControllerRegs[1][0]) #First[] = Row, Second [] = Column

# ----------------------------------------------------- # 

#Global Varaiables
global a
global center_x
global center_y

reg1Val = 0
reg2Val = 0
reg3Val = 0
reg4Val = 0

ControllerIndex = 0 

# ----------------------------------------------------- # 
#Create a data array for the controllers. 
controllerNames = [None] * (numRow-2)
for i in range(0, numRow-2):
    controllerNames[i] = ControllerRegs[i+1][0]
    print(controllerNames[i])
'''
controllerNames = [
    "ABB_ACS550",
    "ABB_ACS580",
    "AID_VFD",
    "Cattron_CP1000",
    "ControlsInc_CSeries",
    "Franklin_PDrive",
    "Fuji_FrenicMega",
    "Lofa_CP750",
    "Lofa_CP750T3",
    "Murhpy_MPC10",
    "Murphy_MPC20",
    "Murphy_TEC10",
    "Toshiba_TE2",
    "WEG_CFW111"
]
'''
# ----------------------------------------------------- # 
#Create an array for used/ polled registers in controller. 
controllerRegisterNames = [
   colList[1], colList[2], colList[3], colList[4]
]
# ----------------------------------------------------- # 
#Create the base JSON file.  
dictionary = {
                "ControllerIndex" : str(ControllerIndex),
                controllerRegisterNames[0] : str(ControllerRegs[ControllerIndex + 1][1]),
                controllerRegisterNames[1] : str(ControllerRegs[ControllerIndex + 1][2]),
                controllerRegisterNames[2] : str(ControllerRegs[ControllerIndex + 1][3]),
                controllerRegisterNames[3] : str(ControllerRegs[ControllerIndex + 1][4]),
                "controllerRegister1Value" : str(reg1Val),
                "controllerRegister2Value" : str(reg2Val),
                "controllerRegister3Value" : str(reg3Val),
                "controllerRegister4Value" : str(reg4Val),
                "lastMessage"              : " "
            }

#Create the json.
#json_object = json.dumps(dictionary, indent = 4)

#Write to the data json. 
with open(jsonFile, 'w') as outfile: 
    json.dump(dictionary, outfile, indent =4)

# ----------------------------------------------------- # 
#Create the base app that will load the individual screens. 

class SampleApp(tk.Tk):

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        # the container is where we'll stack a bunch of frames
        # on top of each other, then the one we want visible
        # will be raised above the others
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        
        #Get the screen dimensions. 
        screen_width =  container.winfo_screenwidth()
        screen_height = container.winfo_screenheight()

        #Set the window dimensions to match the screen. 
        window_width = screen_width
        window_height = screen_height

        #Find the center point.
        center_x = int(screen_width/2 - window_width/2)
        center_y = int(screen_height/2 - window_height/2)
        print(center_x)
        #Set the window geometry. 
        self.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

        #Set the fonts used.  
        self.title_font = tkfont.Font(family='Helvetica', size=30, weight="bold", slant="italic")
        self.button_font = tkfont.Font(family='Helvetica', size=25, weight="bold", slant="italic")
        self.exit_font = tkfont.Font(family='Helvetica', size=40, weight="bold", slant="italic")
        self.fit_font = tkfont.Font(family='Helvetica', size=15, weight="bold", slant="italic")

        self.frames = {}
        for F in (MainMenu, ControllerSelect, ControllerDisplay):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame

            # put all of the pages in the same location;
            # the one on the top of the stacking order
            # will be the one that is visible.
            frame.grid(row=0, column=0, sticky="nsew")
        self.show_frame("MainMenu")

    def show_frame(self, page_name):
        '''Show a frame for the given page name'''
        frame = self.frames[page_name]
        frame.tkraise()

# ----------------------------------------------------- # 
#Create a class for the start page. 

class MainMenu(tk.Frame, SampleApp):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        #Assign Controller Values 
        self.controller = controller
        controller.i = 0
        controller.name = "None"

        #Create Buttons and Labels 
        self.buttonImg = tk.PhotoImage(file = button) 
        #self.buttonImg = self.buttonImg.subsample(2)
        label = tk.Label(self, text="Cornell Spooky", font=controller.title_font)
        label.pack(side="top", fill="x", pady=15)

        button1 =tk.Button(self, text="Advance", font=controller.button_font, image = self.buttonImg, borderwidth=0, compound = "center",
                         command=lambda: controller.show_frame("ControllerSelect")).pack(pady = 15)
        button2 = tk.Button(self, text = "Quit", font=controller.button_font, image = self.buttonImg, borderwidth=0, compound = "center",
                            command=lambda: quit(SampleApp)).pack(pady = 5)
        
    def quit(SampleApp):
        SampleApp.destroy()
        
        #label1 = tk.Label(button1, text="Advance", font=controller.title_font)
        #label2 = tk.Label(button2, text="Quit", font=controller.title_font)

        #label1.pack()
        #label2.pack()
        
        #button1.grid( row = 3, column = 1, columnspan = 2)
        #button2.grid( row = 3, column = 2, columnspan = 2)

        #button1.pack(pady = 30)
        #button2.pack(pady = 80)

# ----------------------------------------------------- # 
#Create a class for the controller selection page. 

class ControllerSelect(tk.Frame):

    def __init__(self, parent, controller):
        
        #Create frame and container info. 
        tk.Frame.__init__(self, parent)
        container = tk.Frame(self)
        self.controller = controller

        #Get the screen dimensions. 
        screen_width =  container.winfo_screenwidth()
        screen_height = container.winfo_screenheight()

        #Set the window dimensions to match the screen. 
        window_width = screen_width
        window_height = screen_height

        #Find the center point.
        center_x = int(screen_width/2)
        center_y = int(screen_height/2)

        #Create labels and buttons. 
        label = tk.Label(self, text="Selected Controller", font=controller.title_font)
        label1 = tk.Label(self, text=controllerNames[0], font=controller.title_font, fg = '#f00')
       
        self.rightArrowImg = tk.PhotoImage(file = rightArrow) 
        self.rightArrowImg = self.rightArrowImg.subsample(2)
        self.leftArrowImg = tk.PhotoImage(file = leftArrow) 
        self.leftArrowImg = self.leftArrowImg.subsample(2)
        self.buttonImg = tk.PhotoImage(file = button)
        #self.buttonImg = self.buttonImg.subsample(2)
        controller.i = 0
        ControllerSelect.varControllerIndex = controller.i


        def inc_label_text(controller): 
            #Update the index value and the label. 
            controller.i = controller.i+1
            if controller.i == len(controllerNames):
                controller.i = 0
            textFill = controllerNames[controller.i]
            label1.config(text = textFill)

            #Predominate Debug Use. -> Loading local variable gives more options, entirely unneccisary. 
            print(controller.i)
            ControllerSelect.varControllerIndex = controller.i
            ControllerSelect.varControllerName = controllerNames[controller.i]
            print(ControllerSelect.varControllerIndex)
            print(ControllerSelect.varControllerName)

            controller.show_frame("ControllerSelect")

            
        def dec_label_text(controller): 
            #Update the index value and the label. 
            controller.i = controller.i-1
            if controller.i == (-1):
                controller.i = len(controllerNames)-1
            textFill = controllerNames[controller.i]
            label1.config(text = textFill)
            self.name = controllerNames[controller.i]

            #Predominate Debug Use. -> Loading local variable gives more options, entirely unneccisary. 
            print(controller.i)
            ControllerSelect.varControllerIndex = controller.i
            ControllerSelect.varControllerName = controllerNames[controller.i]
            print(ControllerSelect.varControllerIndex)
            print(ControllerSelect.varControllerName)

            controller.show_frame("ControllerSelect")
           

        #Create the buttons and labels. 

        button1 = tk.Button(self, image = self.leftArrowImg, borderwidth=0, 
                           command=lambda: dec_label_text(controller))
        
        button2 = tk.Button(self, image = self.rightArrowImg, borderwidth=0, 
                           command=lambda: inc_label_text(controller))
        
        button3 = tk.Button(self, image = self.buttonImg, text = "Select", font=controller.button_font, borderwidth = 0, compound = "center",
                           command=lambda: displayHUD(self))
        
        button_exit = tk.Button(self, text = "X", font=controller.exit_font, borderwidth = 0, 
                           command=lambda: controller.show_frame("MainMenu"))
        button_exit.place(relx = 0.0, rely = 0.0, anchor = 'nw')

        label.pack(side="top", fill="both", pady = 10)
        label1.pack(side="top", fill="both", pady = 30)
        button3.pack(pady = 5)
       
        button1.place(y=center_y, anchor = 'w')
        button2.place(x = screen_width, y=center_y, anchor = 'e')
        #button2.place(y=center_y, anchor = 'e')
       

        def displayHUD(self):
            print("Display the HUD and Load JSON")
            #Create the json dictionary 
            ControllerIndex = ControllerSelect.varControllerIndex
            dictionary2 = {
                "ControllerIndex" : str(ControllerIndex),
                controllerRegisterNames[0] : str(ControllerRegs[ControllerIndex + 1][1]),
                controllerRegisterNames[1] : str(ControllerRegs[ControllerIndex + 1][2]),
                controllerRegisterNames[2] : str(ControllerRegs[ControllerIndex + 1][3]),
                controllerRegisterNames[3] : str(ControllerRegs[ControllerIndex + 1][4]),
                "controllerRegister1Value" : str(reg1Val),
                "controllerRegister2Value" : str(reg2Val),
                "controllerRegister3Value" : str(reg3Val),
                "controllerRegister4Value" : str(reg4Val),
                "lastMessage"              : " "
            }

            

            #Write to the data json. 
            with open(jsonFile, 'w') as outfile: 
                json.dump(dictionary2, outfile, indent =4)
            
            #ControllerDisplay.__init__(self, parent, controller)
            
            #Start the server 
            ServerCommand.startServer() 
            
            #Load the controller display screen. 
        
            controller.show_frame("ControllerDisplay")


# ----------------------------------------------------- # 
#Create a class to display the various register values. 

class ControllerDisplay(tk.Frame):
    
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        object = ControllerSelect(parent, controller)   #This is a redundant object that isn't neccisary. 
        
        self.controller = controller
        #print(controller.i)
        #print(object.varControllerIndex)
        #print(object.varControllerName)

        #Create the buttons and labels. 

        self.uploadImg = tk.PhotoImage(file = upload) 
        self.uploadImg = self.uploadImg.subsample(4)

        self.vertBarImg = tk.PhotoImage(file = vertBar)
        self.vertBarImg = self.vertBarImg.subsample(4)

        label = tk.Label(self, text="Controller : " + controllerNames[0], font=controller.title_font)   #Base label, this is updated using the update button. 
        label.pack(side="top", fill="x", pady=30)

        button_exit = tk.Button(self, text = "X", font=controller.exit_font, borderwidth = 0, 
                            command=lambda: exit(self))
        button_exit.place(relx = 0.00, rely = 0.00, anchor = 'nw')

        button_update = tk.Button(self, image = self.uploadImg, font=controller.title_font, borderwidth = 0, 
                            command=lambda: callServer(self))
        button_update.place(relx = 0.95, rely = 0.05, anchor = 'ne')

        #Graphical Anchor Points 
        barAnchorY = 0.65 
        barAnchorX = [0, 0.25, 0.5, 0.75]

        regTitleAnchorY = 0.3

        regAdressTextAnchorY = 0.4
        regAdressValueAnchorY = 0.45

        x_offset_center = 0.06
        x_offset_indent = 0.02

        regValueTextAnchorY = 0.7
        regValueValueAnchorY = 0.75
        
        statusX = 0
        statusY = 0.9
        
        #Initialize all graphics used.
        
        graphic1 = tk.Label(self, image = self.vertBarImg, borderwidth = 0)
        graphic1.place(relx = barAnchorX[1], rely = barAnchorY, anchor = 'w')

        graphic2 = tk.Label(self, image = self.vertBarImg, borderwidth = 0)
        graphic2.place(relx = barAnchorX[2], rely = barAnchorY, anchor = 'w')

        graphic3 = tk.Label(self, image = self.vertBarImg, borderwidth = 0)
        graphic3.place(relx = barAnchorX[3], rely = barAnchorY, anchor = 'w')

        registerTitle1 = tk.Label(self, text = controllerRegisterNames[0], font=controller.fit_font, borderwidth = 0)
        registerTitle2 = tk.Label(self, text = controllerRegisterNames[1], font=controller.fit_font, borderwidth = 0)
        registerTitle3 = tk.Label(self, text = controllerRegisterNames[2], font=controller.fit_font, borderwidth = 0)
        registerTitle4 = tk.Label(self, text = controllerRegisterNames[3], font=controller.fit_font, borderwidth = 0)

        registerTitle1.place(relx = barAnchorX[0] + x_offset_center, rely = regTitleAnchorY, anchor = 'nw')
        registerTitle2.place(relx = barAnchorX[1] + x_offset_center, rely = regTitleAnchorY, anchor = 'nw')
        registerTitle3.place(relx = barAnchorX[2] + x_offset_center, rely = regTitleAnchorY, anchor = 'nw')
        registerTitle4.place(relx = barAnchorX[3] + x_offset_center, rely = regTitleAnchorY, anchor = 'nw')

        registerAddressText1 = tk.Label(self, text = "Address", font=controller.fit_font, borderwidth = 0)
        registerAddressText2 = tk.Label(self, text = "Address", font=controller.fit_font, borderwidth = 0)
        registerAddressText3 = tk.Label(self, text = "Address", font=controller.fit_font, borderwidth = 0)
        registerAddressText4 = tk.Label(self, text = "Address", font=controller.fit_font, borderwidth = 0)

        registerAddressText1.place(relx = barAnchorX[0] + x_offset_indent, rely = regAdressTextAnchorY, anchor = 'nw')
        registerAddressText2.place(relx = barAnchorX[1] + x_offset_indent, rely = regAdressTextAnchorY, anchor = 'nw')
        registerAddressText3.place(relx = barAnchorX[2] + x_offset_indent, rely = regAdressTextAnchorY, anchor = 'nw')
        registerAddressText4.place(relx = barAnchorX[3] + x_offset_indent, rely = regAdressTextAnchorY, anchor = 'nw')

        registerAddressValue1 = tk.Label(self, text = "0", font=controller.fit_font, borderwidth = 0)
        registerAddressValue2 = tk.Label(self, text = "0", font=controller.fit_font, borderwidth = 0)
        registerAddressValue3 = tk.Label(self, text = "0", font=controller.fit_font, borderwidth = 0)
        registerAddressValue4 = tk.Label(self, text = "0", font=controller.fit_font, borderwidth = 0)

        registerAddressValue1.place(relx = barAnchorX[0] + x_offset_indent, rely = regAdressValueAnchorY, anchor = 'nw')
        registerAddressValue2.place(relx = barAnchorX[1] + x_offset_indent, rely = regAdressValueAnchorY, anchor = 'nw')
        registerAddressValue3.place(relx = barAnchorX[2] + x_offset_indent, rely = regAdressValueAnchorY, anchor = 'nw')
        registerAddressValue4.place(relx = barAnchorX[3] + x_offset_indent, rely = regAdressValueAnchorY, anchor = 'nw')

        registerValueText1 = tk.Label(self, text = "Value", font=controller.fit_font, borderwidth = 0)
        registerValueText2 = tk.Label(self, text = "Value", font=controller.fit_font, borderwidth = 0)
        registerValueText3 = tk.Label(self, text = "Value", font=controller.fit_font, borderwidth = 0)
        registerValueText4 = tk.Label(self, text = "Value", font=controller.fit_font, borderwidth = 0)

        registerValueText1.place(relx = barAnchorX[0] + x_offset_indent, rely = regValueTextAnchorY, anchor = 'nw')
        registerValueText2.place(relx = barAnchorX[1] + x_offset_indent, rely = regValueTextAnchorY, anchor = 'nw')
        registerValueText3.place(relx = barAnchorX[2] + x_offset_indent, rely = regValueTextAnchorY, anchor = 'nw')
        registerValueText4.place(relx = barAnchorX[3] + x_offset_indent, rely = regValueTextAnchorY, anchor = 'nw')

        registerValueValue1 = tk.Label(self, text = "--", font=controller.fit_font, borderwidth = 0)
        registerValueValue2 = tk.Label(self, text = "--", font=controller.fit_font, borderwidth = 0)
        registerValueValue3 = tk.Label(self, text = "--", font=controller.fit_font, borderwidth = 0)
        registerValueValue4 = tk.Label(self, text = "--", font=controller.fit_font, borderwidth = 0)

        registerValueValue1.place(relx = barAnchorX[0] + x_offset_indent, rely = regValueValueAnchorY, anchor = 'nw')
        registerValueValue2.place(relx = barAnchorX[1] + x_offset_indent, rely = regValueValueAnchorY, anchor = 'nw')
        registerValueValue3.place(relx = barAnchorX[2] + x_offset_indent, rely = regValueValueAnchorY, anchor = 'nw')
        registerValueValue4.place(relx = barAnchorX[3] + x_offset_indent, rely = regValueValueAnchorY, anchor = 'nw')
        
        statusMessage = tk.Label(self, text = "Awaiting Server Connection", font = controller.fit_font, borderwidth = 0)
        #statusMessage.place(relx = statusX, relyy = statusY, anchor = 'nw')
        statusMessage.pack(side = "bottom", fill = "x", pady = 20)

        def exit(self):
            
            ServerCommand.killServer()
            controller.show_frame("ControllerSelect")


        def labelUpdate(self):
            #threading.Thread(target=labelUpdate)
            with open(jsonFile, 'r') as openfile:
                json_object = json.load(openfile)

            print("This is what I SENT")
            print(json_object)

            dictionary = json_object 

            print(dictionary)

            controllerIndex = int(dictionary['ControllerIndex'])
            textFill = controllerNames[int(controllerIndex)]
            label.config(text = "Controller : " + textFill)

            registerAddressValue1.config(text = dictionary[str(controllerRegisterNames[0])])
            registerAddressValue2.config(text = dictionary[str(controllerRegisterNames[1])])
            registerAddressValue3.config(text = dictionary[str(controllerRegisterNames[2])])
            registerAddressValue4.config(text = dictionary[str(controllerRegisterNames[3])])
            
            print("Labels updated sucsessfully")
            
            statusMessage.config(text = "Values updated, program initialized...")
            registerValueValue1.config(text = dictionary["controllerRegister1Value"])
            registerValueValue2.config(text = dictionary["controllerRegister2Value"])
            registerValueValue3.config(text = dictionary["controllerRegister3Value"])
            registerValueValue4.config(text = dictionary["controllerRegister4Value"])
            
           
           
            
        def callServer(self):
              #START THE SERVER LAST
            labelUpdate(self)
            #ServerCommand.startServer()

           
        #labelUpdate(self)
        #threading.Timer(10, labelUpdate(self)).start()
    #Implement threading on the labelUpdate function. 
        #threading.Timer(2, labelUpdate(self)).start()
        
# ----------------------------------------------------- # 
# Create a class to handle the starting/ ending of the server. 

class ServerCommand():
    signal = False 
    
    def startServer(): 
        subprocess.run("python3 " + serverCode + " &", shell = True)
        print("Server Started")
        signal = True 
        
    def killServer():
        serverPID = str(ServerCommand.getPID("TGODServer"))
        print("I got the serverPID " + serverPID)  
        os.system("sudo kill -9 " + serverPID)
        print("Server Terminated Sucsessfully")
        signal = False
        
    def getPID(name): 
        return int(check_output(["pidof","-s",name]))
        
    
#This will run the main loop. 
if __name__ == "__main__":
    app = SampleApp()
    app.mainloop()
# ----------------------------------------------------- # 
